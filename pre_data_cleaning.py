# -*-coding:utf-8-*-
# previous thing to do: data cleaning
# convert tmp_data_label.xlsx into tmp_data_label_cleaned.xlsx

import os 
import openpyxl
import re

# definitions
label_dir = []
label_file = []
label_num = []
label_proportion = []
dir_count = 0
file_count = 0

# functions & definitions

def clean(tmp_str):
    tmp_str = tmp_str.lower() # lowercase
    tmp_str = re.sub("[^a-z-]+", " ", tmp_str) # remove punctuations
    tmp_str = tmp_str.strip() # remove extra space
    return tmp_str

from nltk.stem import SnowballStemmer
def stem_snowball(str):  
    str_list = str.split(" ")
    for word in str_list:
        stemmer = SnowballStemmer("german")
        word = stemmer.stem(word)
    return ' '.join(str_list)

# parameters
sen_num = 5042 # sen: index from 0 to 5042
stem = lambda x: x
# stem = stem_snowball


wb = openpyxl.load_workbook('tmp_data_label.xlsx')
sheet = wb.active
wb1 = openpyxl.Workbook()
ws = wb1.active
for i in range(sen_num):
    # data_str: remove punctuations & stem
    tmp_str = str(sheet.cell(row = i+2, column = 1).value)
    tmp_str = clean(tmp_str)
    tmp_str = stem(tmp_str)
    ws.cell(row = i+2, column = 1).value = tmp_str
    # label_dir, file
    if sheet.cell(row = i+2, column = 2).value != sheet.cell(row = i+1, column = 2).value:
        dir_count = dir_count + 1
    if sheet.cell(row = i+2, column = 3).value != sheet.cell(row = i+1, column = 3).value:
        file_count = file_count + 1
    ws.cell(row = i+2, column = 2).value = dir_count
    ws.cell(row = i+2, column = 3).value = file_count
    # label_num: used later
    label_dir.append(str(sheet.cell(row = i+2, column = 2).value))
    label_file.append(str(sheet.cell(row = i+2, column = 3).value))
    label_num.append(sheet.cell(row = i+2, column = 4).value)
    # move, step, ms: unchange
    ws.cell(row = i+2, column = 5).value = str(sheet.cell(row = i+2, column = 5).value)
    ws.cell(row = i+2, column = 6).value = str(sheet.cell(row = i+2, column = 6).value)
    ws.cell(row = i+2, column = 7).value = str(sheet.cell(row = i+2, column = 7).value)

label_file_setlist = list(set(label_file))
label_dir_setlist = list(set(label_dir))
# dir
'''for j in range(len(label_dir_setlist)):
    dir1 = label_dir_setlist[j]
    tmp_index = [i for i in range(len(label_dir)) if label_dir[i] == dir1]
    for i in tmp_index:
        ws.cell(row = i+2, column = 2).value = j'''

# file, proportion
for j in range(len(label_file_setlist)):
    file = label_file_setlist[j]
    tmp_index = [i for i in range(len(label_file)) if label_file[i] == file]
    tmp_max_sentence = max([int(label_num[i]) for i in tmp_index])
    for i in tmp_index:
        # ws.cell(row = i+2, column = 3).value = j
        ws.cell(row = i+2, column = 4).value = (float(label_num[i]) - 2) / (tmp_max_sentence - 2)
        # label_proportion[i] = (label_proportion[i] - 2) / (tmp_max_sentence - 2)

# last work
ws.cell(row = 1, column = 1).value = 'data_str_cleaned'
ws.cell(row = 1, column = 2).value = 'label_dir_serialnum'
ws.cell(row = 1, column = 3).value = 'label_file_serialnum'
ws.cell(row = 1, column = 4).value = 'label_proportion'
ws.cell(row = 1, column = 5).value = 'label_move'
ws.cell(row = 1, column = 6).value = 'label_step'
ws.cell(row = 1, column = 7).value = 'label_movestep'
wb1.save('tmp_data_label_cleaned.xlsx')
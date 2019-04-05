# -*-coding:utf-8-*-
import os 
import openpyxl
import re
import numpy as np
import time
import math

from func_vectorize import *
# from func_vectorize import vectorize
from func_cv import *
# from func_cv_multi_thread import *
# from func_cv import cv
from func_evaluation import *
# from func_evaluation import cv

print("Start Processing")
print("Local Time:", time.asctime(time.localtime(time.time())))

# definition
data_str = [] # the original sentence
data_list = [] # each sentence is represented as a list
data_vector = [] # sentence vector + additional features
label_dir = [] # belong to which dir
label_file = [] # belong to which file
label_proportion = [] # sentence in file: front or back? 0~1
label_move = []
label_step = []
label_movestep = []

# para: normal
eps = 1e-3
sen_num = 1000
# sen_num = 5042 # sen: index from 0 to 5042

# para: vactorization
data_vector_judge = False # True: no judge; False: judge by returning result of vectorization

# para: undefined detection: not used
# undefined = 0
# undefined_threshold = 0.5

# para: cv
# label_batch = label_file
batch_num = 10

# functions
def batch_generate(label_num, batch_num):
    # generate random labels of batch_num types
    return [i % batch_num + 1 for i in range(label_num)]

attrs = ['data_str_cleaned', 'label_dir_serialnum', 'label_file_serialnum', 'label_proportion', 
         'label_move', 'label_step', 'label_movestep'] # just to mention

# data prepare
wb = openpyxl.load_workbook('../data/tmp_data_label_cleaned.xlsx')
sheet = wb.active
except_num = 0
for i in range(sen_num):
    data_str_elem = str(sheet.cell(row = i+2, column = 1).value)
    [data_vector_elem, data_vector_ok, data_vector_word_count] = vectorize(data_str_elem)
    if data_vector_judge or data_vector_ok:
        data_str.append(data_str_elem)
        data_vector.append(data_vector_elem)
        label_dir.append(str(sheet.cell(row = i+2, column = 2).value))
        label_file.append(str(sheet.cell(row = i+2, column = 3).value))
        label_proportion.append(str(sheet.cell(row = i+2, column = 4).value))
        label_move.append(str(sheet.cell(row = i+2, column = 5).value))
        label_step.append(str(sheet.cell(row = i+2, column = 6).value))
        label_movestep.append(str(sheet.cell(row = i+2, column = 7).value))
    else:
        except_num = except_num + 1

print(except_num, "Out of", sen_num, "Samples is Abandoned, Percentage =", format(except_num / sen_num * 100,'.2f'))

label_batch = batch_generate(sen_num - except_num, batch_num)

print("Before CV, time =", time.process_time(), "seconds")
[test_confuse_matrix, train_confuse_matrix] = cv(data_vector, label_move, label_batch)




